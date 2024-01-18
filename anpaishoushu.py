import os
import sys
import easygui as E
import xlrd
import openpyxl
import datetime
import calendar
import re
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
num_in_each_room = [0 for i in range(31)]
col_index = [0 for i in range(31)]
file_path = ''
wb = None
date = datetime.datetime.now().date()
# WPSid=345470524
# wps_path = os.environ['USERPROFILE']+'\\Documents\\WPSDrive\\'+str(WPSid)+'\\WPS云盘\\共享文~1\\群共享文件\\'
wps_path = ''

def data_collect(book):
    type_list = []
    name_list = []
    sh=book.sheet_by_index(0)
    for i in range(sh.nrows):
        type_list.append(sh.cell_value(i,0))
    for i in range(len(type_list)):
        j=1
        tmp=[]
        for j in range(1,sh.ncols):
            if sh.cell_type(i,j)==0:
                continue
            tmp.append(sh.cell_value(i,j))
        name_list.append(tmp)
    return (type_list,name_list)

def del_lines(ws):
    #删除原表中多余信息
    # ws.page_setup.fitToHeight = 1
    ws.unmerge_cells('A1:Q1')
    ws.delete_cols(14)#巡回护士
    ws.delete_cols(13)#洗手护士
    ws.delete_cols(12)#手术助手
    ws.delete_cols(10)#备注
    for i in range(1, ws.max_row+1):
        ws.row_dimensions[i].height = 0
    for row in range(ws.max_row,2,-1):#最后一行开始删除局麻
        type = ws.cell(row = row, column = ws.max_column).value
        # print(type)
        if type=='局麻' or type == '局部麻醉':
            ws.delete_rows(row)
    colors=['E2EFDA','FFF2CC','E6E6FA']
    vis_name = set()
    for index,row in enumerate(ws.rows):
        if index <= 1:
            continue
        row[10].font=Font('仿宋',14,True)
        row[11].font=Font('仿宋',14,True)
        row[10].alignment=Alignment('center','center')
        row[11].alignment=Alignment('center','center')
        room_name=row[0].value
        # 分隔线
        if room_name not in vis_name:
            for each in row:
                each.border=border = Border(left=Side(border_style='thin',color='000000'), \
                    right=Side(border_style='thin',color='000000'), \
                    top=Side(border_style='thick',color='000000'), \
                    bottom=Side(border_style='thin',color='000000'))
        vis_name.add(room_name)
        row[0].fill=PatternFill('solid',fgColor=colors[len(vis_name)%2])
        # 年龄警示
        age = row[6].value
        num = int("".join(list(filter(str.isdigit,age))))
        if '月' in age or num<=3 or num>=80:
            row[6].fill=PatternFill('solid',fgColor='ff0000')
        # 区分心脏
        # flag=0
        # for i,each in enumerate(room_partition):
        #     if room_name in each:
        #         flag=1
        #         for cell in row:
        #             cell.fill=PatternFill('solid',fgColor=colors[i%2])
        #         break
        # if flag==0:
        #     for cell in row:
        #         cell.fill=PatternFill('solid',fgColor=colors[2])

    string = ws['A1'].value
    string += '  麻醉总数:' + str(ws.max_row-2)
    ws['A1'] = string
    ws.merge_cells('A1:M1')

    ws.column_dimensions['K'].width=20
    ws.column_dimensions['L'].width=20

    for i in range(1, ws.max_row+1):
        ws.row_dimensions[i].height = 28

# def relieve():
#     pass

def write_cell(ws,x,y,s,fn=u'宋体',fs=14):
    ws.cell(x,y).font=Font(name=fn, size=fs)
    ws.cell(x,y).alignment=Alignment('center','center')
    ws.cell(x,y).value=s

def get_oneday(sh,delta,type_list) -> list:
    global date
    oneday = date + datetime.timedelta(days=delta)
    date_list = calendar.monthcalendar(oneday.year,oneday.month)
    tmp = []
    if oneday.year==date.year and oneday.month==date.month:
        pass
    else:
        xls_name = str(oneday.year)+'年'+str(oneday.month)+'月值班表.xls'
        path_pb=wps_path+'\\群共享文件\\'+str(oneday.year)+'年值班表\\'+xls_name
        if not os.path.isfile(path_pb):
            print(path_pb)
            E.msgbox(msg='需要采集{}年{}月信息，找不到当月排班表！'.format(oneday.year,oneday.month))
            return []
        sh=xlrd.open_workbook(path_pb).sheet_by_index(0)
    for i in range(len(date_list)):
        for j in range(7):
            if date_list[i][j]==oneday.day:
                for k in range(len(type_list)):
                    tmp.append(sh.cell_value(i*(len(type_list)+1)+k+1+1+1,j+1))
    return tmp

def avai_nextday(ws):
    global date
    if not os.path.isfile('data.xls'):
        E.msgbox(msg='没有准备数据库文件！（data.xls）')
        sys.exit()
    type_list,name_list = data_collect(xlrd.open_workbook('data.xls',formatting_info=True))
    doc_dict = dict(zip(type_list, name_list))
    # print(doc_dict)
    date_list = calendar.monthcalendar(date.year,date.month)

    avai_doc=doc_dict['三线']+doc_dict['二线（白）']+doc_dict['一线']
    print(avai_doc)
    xls_name = str(date.year)+'年'+str(date.month)+'月值班表.xls'
    path_pb=wps_path+'\\群共享文件\\'+str(date.year)+'年值班表\\'+xls_name
    if not os.path.isfile(path_pb):
        print(path_pb)
        E.msgbox(msg='找不到当月排班表！')
        return
    #去除下夜班人员
    sh1 = xlrd.open_workbook(path_pb).sheet_by_index(0)
    hfs = get_oneday(sh1,-2,type_list)#for 恢复室
    xyb = get_oneday(sh1,-1,type_list)#下夜班
    zb = get_oneday(sh1,0,type_list)#值班
    bb = get_oneday(sh1,2,type_list)#备班
    print(xyb)
    for i in xyb:
        if i in avai_doc:
            avai_doc.remove(i)
    print(avai_doc)
    tmp=[]
    for i in avai_doc:
        tmp.append("".join(i.split()))
    avai_doc=tmp
    print(avai_doc)
    #去除请假人员
    qj = []
    # date = datetime.datetime.now().date()
    path_qj=wps_path+'\\群共享文件\\'+str(date.year)+'年请假表\\'+str(date.year)+'年'+str(date.month)+'月请假表.xls'
    if not os.path.isfile(path_qj):
        print(path_qj)
        E.msgbox(msg='找不到当月请假表！')
    else:
        sh0 = xlrd.open_workbook(path_qj).sheet_by_index(0)
        height = 10
        for i in range(len(date_list)):
            for j in range(7):
                if date_list[i][j]==date.day:#请假
                    for k in range(1,10):
                        try:
                            qj.append(sh0.cell_value(i*(height+1)+2+k,j))
                        except:
                            continue
        print(qj)
        qj = list(filter(lambda x:x, qj))
        for i in qj:
            if i in avai_doc:
                avai_doc.remove(i)
    #打印信息
    basx = ws.max_row+2
    basy= ws.max_column-4
    ws.cell(basx,basy).value=date.strftime('%Y/%m/%d')
    ws.cell(basx,basy).font=Font(name=u'宋体', size=14, bold=True, color='FF0000')
    write_cell(ws,basx+2,basy+2,(date-datetime.timedelta(2)).day)
    write_cell(ws,basx+2,basy+3,(date-datetime.timedelta(1)).day)
    write_cell(ws,basx+2,basy+4,date.day)
    write_cell(ws,basx+2,basy+5,(date+datetime.timedelta(2)).day)
    # write_cell(ws,basx+2,basy+6,'通道1')
    # write_cell(ws,basx+2,basy+7,'通道2')
    #胃镜恢复室
    # write_cell(ws,basx+3,basy+6,hfs[2],fn=u'楷体')
    # write_cell(ws,basx+3,basy+8,hfs[3],fn=u'楷体')
    data = {'请假':qj+['---']+avai_doc,' ':type_list,'  ':hfs,'下夜班':xyb,'值班':zb,'备班':bb,'麻醉门诊及胃镜':[],'恢复室':[]}
    tmp=[]
    for i in zb:
        tmp.append("".join(i.split()))
    zb=tmp
    print(zb)
    for i,item in enumerate(data.keys()):
        ws.cell(basx+1,basy+i).value=item
        ws.cell(basx+1,basy+i).font=Font(name=u'宋体', size=14)
        ws.cell(basx+1,basy+i).alignment=Alignment('center','center')
    for i,l in enumerate(data.values()):
        for j in range(len(l)):
            ws.cell(basx+3+j,basy+i).value=l[j]
            ws.cell(basx+3+j,basy+i).font=Font(name=u'楷体', size=14)
            ws.cell(basx+3+j,basy+i).alignment=Alignment('center','center')
            if i==0 and l[j] in zb:
                ws.cell(basx+3+j,basy+i).font = Font(name=u'楷体', size=14, color='006400')
    for i in range(basx, ws.max_row+1):
        ws.row_dimensions[i].height = 17.5
    # ws.column_dimensions["O"].weight = 12
    # ws.column_dimensions["P"].weight = 12
    # ws.merge_cells('O{}:P{}'.format(basx+1,basx+1))

def open_file():
    global file_path,wb,date,wps_path
    # file_path = E.fileopenbox(title='打开文件',msg='打开要填写的表格',default=os.environ['USERPROFILE']+'\\Documents\\WPSDrive\\'+str(WPSid)+'\\WPS云盘\\共享文~1\\手术排班共享\\*.xlsx')
    file_path = E.fileopenbox(title='打开文件',msg='打开要填写的表格')
    if file_path==None:
        sys.exit()
    wps_path = os.path.dirname(os.path.dirname(os.path.dirname(file_path)))
    print(wps_path)
    try:
        date = datetime.datetime.strptime(file_path[-16:-6],'%Y-%m-%d').date()
    except:
        E.msgbox(msg='无法解析表格时间，请检查是否选错文件')
        sys.exit()
    print(date)
    wb = openpyxl.load_workbook(file_path)

def save_file():
    global file_path,wb
    new_str=file_path[:-5] + '(new).xlsx'
    flag=1
    while flag:
        try:
            wb.save(new_str)
            flag=0
        except PermissionError as e:
            r=E.msgbox(msg=repr(e)+'\n\n文件可能正在被其他应用或其他人占用！')
            if r is None:
                return
        except FileNotFoundError as e:
            r=E.msgbox(msg=repr(e)+'\n\n文件保存路径路径错误！')
            return
        except Exception as e:
            r=E.msgbox(msg=repr(e))
            return
    os.startfile(new_str)

def main():
    global wb
    # avai_nextday()
    open_file()
    del_lines(wb.active)
    avai_nextday(wb.active)
    save_file()
    # #init
    # data_path = E.fileopenbox(title='打开文件',msg='请选择数据文件',default='*.xls')
    # book = xlrd.open_workbook(data_path)
    # assistant = Assistant.import_all(book.sheet_by_index(2))
    # doctor    =    Doctor.import_all(book.sheet_by_index(1))

    # filling_blanks(assistant,doctor)

if(__name__=='__main__'):
    main()